#!/usr/bin/env python3
"""
CyberArk Safe AD Group "Rename" Batch Tool
------------------------------------------
Implements "rename" as:
  1) GET old safe member (group) to capture permissions + expiration
  2) POST new safe member with same permissions + expiration
     - or PUT update if new already exists (optional)
  3) DELETE old safe member (optional)

Docs (PAM Self-Hosted):
- Logon:  /PasswordVault/API/auth/{Cyberark|LDAP|Windows|RADIUS}/Logon/  (POST) :contentReference[oaicite:1]{index=1}
- Get member: /PasswordVault/API/Safes/{SafeUrlId}/Members/{MemberName}/ (GET) :contentReference[oaicite:2]{index=2}
- Add member: /PasswordVault/API/Safes/{SafeUrlId}/Members/             (POST) :contentReference[oaicite:3]{index=3}
- Update member: /PasswordVault/API/Safes/{SafeUrlId}/Members/{MemberName}/ (PUT) :contentReference[oaicite:4]{index=4}
- Delete member: /PasswordVault/API/Safes/{SafeUrlId}/Members/{MemberName}/ (DELETE) :contentReference[oaicite:5]{index=5}
"""

from __future__ import annotations

import argparse
import concurrent.futures as futures
import dataclasses
import datetime as dt
import json
import logging
import os
import re
import sys
import time
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote

import requests

try:
    import openpyxl  # type: ignore
except ImportError:
    openpyxl = None  # noqa


__version__ = "0.1.0"


# ---------------------------
# Data model
# ---------------------------

@dataclasses.dataclass(frozen=True)
class RenameOp:
    environment: str
    safe_name: str
    old_group: str
    new_group: str
    # Optional columns you mentioned (kept for reporting; not CyberArk API fields):
    costco_object_approver: Optional[Any] = None
    costco_object_author_sme: Optional[Any] = None
    # Row number for better diagnostics:
    rownum: Optional[int] = None


# ---------------------------
# Logging
# ---------------------------

def setup_logging(level: str, mode: str, logfile: Optional[str]) -> None:
    lvl = getattr(logging, level.upper(), None)
    if not isinstance(lvl, int):
        raise ValueError(f"Invalid loglevel: {level}")

    handlers: List[logging.Handler] = []
    if mode == "stdout":
        handlers.append(logging.StreamHandler(sys.stdout))
    elif mode == "file":
        if not logfile:
            raise ValueError("--logfile is required when --logmode=file")
        handlers.append(logging.FileHandler(logfile, encoding="utf-8"))
    else:
        raise ValueError(f"Invalid logmode: {mode}")

    logging.basicConfig(
        level=lvl,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=handlers,
    )


# ---------------------------
# Excel parsing (flexible headers)
# ---------------------------

def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", "", s)
    s = s.replace("_", "").replace("-", "")
    return s


def _get_cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def read_ops_from_xlsx(path: str, sheet: str, env_filter: Optional[str]) -> List[RenameOp]:
    if openpyxl is None:
        raise RuntimeError("Missing dependency: openpyxl. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = rows[0]
    idx: Dict[str, int] = {}
    for i, h in enumerate(header):
        idx[_norm_header(h)] = i

    def pick(*candidates: str) -> Optional[int]:
        for c in candidates:
            key = _norm_header(c)
            if key in idx:
                return idx[key]
        return None

    # Expected columns (very tolerant):
    i_env = pick("Environment", "Env")
    i_safe = pick("SafeName", "Safe", "SafeUrlId")
    i_old = pick("ExistingSecurityGroupName", "ExistingGroupName", "OldGroup", "Existing", "CurrentGroup")
    i_new = pick("NewSecurityGroupName", "NewGroupName", "NewGroup", "New", "TargetGroup")

    # Optional custom cols you mentioned:
    i_approver = pick("costco-object-approver", "costcoobjectapprover", "objectapprover")
    i_author = pick("costco-object-author-sme", "costcoobjectauthorsme", "objectauthorsme")

    missing = []
    if i_env is None: missing.append("Environment")
    if i_safe is None: missing.append("SafeName")
    if i_old is None: missing.append("Existing group name")
    if i_new is None: missing.append("New group name")
    if missing:
        raise ValueError(f"Missing required column(s) in '{sheet}': {', '.join(missing)}")

    ops: List[RenameOp] = []
    for r_i, r in enumerate(rows[1:], start=2):
        env = _get_cell_str(r[i_env])  # type: ignore[index]
        safe = _get_cell_str(r[i_safe])  # type: ignore[index]
        oldg = _get_cell_str(r[i_old])  # type: ignore[index]
        newg = _get_cell_str(r[i_new])  # type: ignore[index]
        if not (env and safe and oldg and newg):
            continue
        if env_filter and env != env_filter:
            continue

        ops.append(
            RenameOp(
                environment=env,
                safe_name=safe,
                old_group=oldg,
                new_group=newg,
                costco_object_approver=(r[i_approver] if i_approver is not None else None),
                costco_object_author_sme=(r[i_author] if i_author is not None else None),
                rownum=r_i,
            )
        )
    return ops


def print_info(ops: List[RenameOp]) -> None:
    total = len(ops)
    envs = sorted({o.environment for o in ops})
    safes = len({(o.environment, o.safe_name) for o in ops})
    pairs = len({(o.environment, o.safe_name, o.old_group, o.new_group) for o in ops})

    print(f"Ops: {total}")
    print(f"Environments: {envs}")
    print(f"Unique (env,safe): {safes}")
    print(f"Unique (env,safe,old,new): {pairs}")

    # detect collisions: multiple olds mapping to same new in same safe
    collision: Dict[Tuple[str, str, str], List[RenameOp]] = {}
    for o in ops:
        k = (o.environment, o.safe_name, o.new_group)
        collision.setdefault(k, []).append(o)
    bad = [(k, v) for k, v in collision.items() if len(v) > 1]
    if bad:
        print("\nWARNING: multiple old groups map to same new group inside same (env,safe):")
        for (env, safe, newg), lst in bad[:50]:
            olds = ", ".join(sorted({x.old_group for x in lst}))
            print(f"  - env={env} safe={safe} new={newg} <- olds: {olds}")


# ---------------------------
# CyberArk REST client
# ---------------------------

class CyberArkClient:
    def __init__(
        self,
        base_url: str,
        token: Optional[str],
        verify_tls: bool,
        timeout_s: float,
        retries: int,
        retry_backoff_s: float,
        retry_status: Iterable[int],
        pool_maxsize: int,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.verify_tls = verify_tls
        self.timeout_s = timeout_s
        self.retries = retries
        self.retry_backoff_s = retry_backoff_s
        self.retry_status = set(retry_status)

        self.session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(pool_connections=pool_maxsize, pool_maxsize=pool_maxsize)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

        self.token = token

    def _headers(self) -> Dict[str, str]:
        h = {"Content-Type": "application/json"}
        if self.token:
            h["Authorization"] = self.token
        return h

    def _req(self, method: str, url: str, **kwargs: Any) -> requests.Response:
        last_exc: Optional[BaseException] = None
        for attempt in range(self.retries + 1):
            try:
                resp = self.session.request(
                    method,
                    url,
                    headers={**self._headers(), **kwargs.pop("headers", {})},
                    timeout=self.timeout_s,
                    verify=self.verify_tls,
                    **kwargs,
                )
                if resp.status_code in self.retry_status and attempt < self.retries:
                    sleep = self.retry_backoff_s * (2 ** attempt)
                    logging.warning("HTTP %s %s -> %s (retry in %.2fs)", method, url, resp.status_code, sleep)
                    time.sleep(sleep)
                    continue
                return resp
            except requests.RequestException as e:
                last_exc = e
                if attempt < self.retries:
                    sleep = self.retry_backoff_s * (2 ** attempt)
                    logging.warning("HTTP %s %s failed: %s (retry in %.2fs)", method, url, e, sleep)
                    time.sleep(sleep)
                    continue
                raise
        if last_exc:
            raise last_exc
        raise RuntimeError("unreachable")

    def logon(self, auth_type: str, username: str, password: str, concurrent_session: bool) -> str:
        url = f"{self.base_url}/PasswordVault/API/auth/{auth_type}/Logon/"
        body: Dict[str, Any] = {
            "username": username,
            "password": password,
        }
        if concurrent_session:
            body["concurrentSession"] = True

        resp = self._req("POST", url, json=body)   # <-- use json=
        if resp.status_code >= 300:
            raise RuntimeError(f"Logon failed: HTTP {resp.status_code}: {resp.text}")

        token = resp.text.strip().strip('"')
        self.token = token
        return token

    def logoff(self) -> None:
        url = f"{self.base_url}/PasswordVault/API/Auth/Logoff/"
        resp = self._req("POST", url)
        if resp.status_code not in (200, 204):
            logging.warning("Logoff returned HTTP %s: %s", resp.status_code, resp.text)

    @staticmethod
    def _safe_url_id(safe_name: str) -> str:
        return quote(safe_name, safe="")

    @staticmethod
    def _member_name(member: str) -> str:
        return quote(member, safe="")

    def get_member(self, safe: str, member: str) -> Optional[Dict[str, Any]]:
        # :contentReference[oaicite:7]{index=7}
        url = f"{self.base_url}/PasswordVault/API/Safes/{self._safe_url_id(safe)}/Members/{self._member_name(member)}/"
        resp = self._req("GET", url)
        if resp.status_code == 404:
            return None
        if resp.status_code >= 300:
            raise RuntimeError(f"GET member failed (safe={safe} member={member}): HTTP {resp.status_code}: {resp.text}")
        return resp.json()

    def add_member(
        self,
        safe: str,
        member: str,
        permissions: Dict[str, Any],
        search_in: str,
        member_type: str,
        membership_expiration_date: Any,
    ) -> Dict[str, Any]:
        # :contentReference[oaicite:8]{index=8}
        url = f"{self.base_url}/PasswordVault/API/Safes/{self._safe_url_id(safe)}/Members/"
        body: Dict[str, Any] = {
            "memberName": member,
            "searchIn": search_in,
            "permissions": permissions,
            "MemberType": member_type,
        }
        # keep expiration only if not empty
        if membership_expiration_date not in (None, "", 0):
            body["membershipExpirationDate"] = membership_expiration_date

        resp = self._req("POST", url, data=json.dumps(body))
        if resp.status_code >= 300:
            raise RuntimeError(f"ADD member failed (safe={safe} member={member}): HTTP {resp.status_code}: {resp.text}")
        return resp.json() if resp.text.strip() else {"status": "ok"}

    def update_member(
        self,
        safe: str,
        member: str,
        permissions: Dict[str, Any],
        membership_expiration_date: Any,
    ) -> Dict[str, Any]:
        # :contentReference[oaicite:9]{index=9}
        url = f"{self.base_url}/PasswordVault/API/Safes/{self._safe_url_id(safe)}/Members/{self._member_name(member)}/"
        body: Dict[str, Any] = {"permissions": permissions}
        if membership_expiration_date not in (None, "", 0):
            body["membershipExpirationDate"] = membership_expiration_date

        resp = self._req("PUT", url, data=json.dumps(body))
        if resp.status_code >= 300:
            raise RuntimeError(f"UPDATE member failed (safe={safe} member={member}): HTTP {resp.status_code}: {resp.text}")
        return resp.json() if resp.text.strip() else {"status": "ok"}

    def delete_member(self, safe: str, member: str) -> None:
        # :contentReference[oaicite:10]{index=10}
        url = f"{self.base_url}/PasswordVault/API/Safes/{self._safe_url_id(safe)}/Members/{self._member_name(member)}/"
        resp = self._req("DELETE", url)
        if resp.status_code in (200, 204):
            return
        if resp.status_code == 404:
            return
        raise RuntimeError(f"DELETE member failed (safe={safe} member={member}): HTTP {resp.status_code}: {resp.text}")


def build_curl_script(
    ops_results: List[Tuple[RenameOp, Dict[str, Any]]],
    base_url: str,
    auth_type: str,
    username: str,
    search_in: str,
    member_type: str,
    delete_old: bool,
) -> str:
    """
    Generates a bash script with curl commands using already-fetched permissions/expiration.
    This avoids dynamic JSON parsing in bash, and is meant for auditable execution.
    """
    lines: List[str] = []
    lines.append("#!/usr/bin/env bash")
    lines.append("set -euo pipefail")
    lines.append("")
    lines.append(f'PVWA_BASE="{base_url.rstrip("/")}"')
    lines.append('TOKEN="${TOKEN:-}"')
    lines.append("")
    lines.append('if [[ -z "$TOKEN" ]]; then')
    lines.append('  echo "TOKEN is empty. Export TOKEN first, or login manually and export it." >&2')
    lines.append('  echo "Example:" >&2')
    lines.append(
        f'  echo "  curl -sk -X POST \\"$PVWA_BASE/PasswordVault/API/auth/{auth_type}/Logon/\\" \\\\" >&2'
    )
    lines.append('  echo "    -H \\"Content-Type: application/json\\" \\\\" >&2')
    lines.append(
        f'  echo "    -d \'{{\\"username\\":\\"{username}\\",\\"password\\":\\"***\\"}}\'" >&2'
    )
    lines.append("  exit 2")
    lines.append("fi")
    lines.append("")

    for op, payload in ops_results:
        safe_q = quote(op.safe_name, safe="")
        old_q = quote(op.old_group, safe="")

        add_body = {
            "memberName": op.new_group,
            "searchIn": search_in,
            "permissions": payload["permissions"],
            "MemberType": member_type,
        }
        if payload.get("membershipExpirationDate") not in (None, "", 0):
            add_body["membershipExpirationDate"] = payload["membershipExpirationDate"]

        # JSON will be embedded in single quotes in bash: -d '...'
        # So we must escape any single quotes inside the JSON for bash safety.
        payload_json = json.dumps(add_body, separators=(",", ":"))
        payload_json = payload_json.replace("'", "'\"'\"'")

        lines.append(f'echo "==> {op.safe_name}: {op.old_group} -> {op.new_group}"')
        lines.append(
            "curl -sS -k -X POST "
            f"\"$PVWA_BASE/PasswordVault/API/Safes/{safe_q}/Members/\" "
            "-H \"Content-Type: application/json\" "
            "-H \"Authorization: ${TOKEN}\" "
            f"-d '{payload_json}'"
        )

        if delete_old:
            lines.append(
                "curl -sS -k -X DELETE "
                f"\"$PVWA_BASE/PasswordVault/API/Safes/{safe_q}/Members/{old_q}/\" "
                "-H \"Authorization: ${TOKEN}\""
            )

        lines.append("")

    return "\n".join(lines)


def run_one_op(
    op: RenameOp,
    client: CyberArkClient,
    search_in: str,
    member_type: str,
    on_conflict: str,
    delete_old: bool,
    dry_run: bool,
    allow_missing_old: bool,
    sleep_s: float,
) -> Tuple[RenameOp, str, Optional[Dict[str, Any]]]:
    """
    Returns: (op, status, captured_old_member_payload_for_curl_or_audit)
    """
    if sleep_s > 0:
        time.sleep(sleep_s)

    if op.old_group == op.new_group:
        return (op, "SKIP_SAME_NAME", None)

    old_member = client.get_member(op.safe_name, op.old_group)
    if old_member is None:
        msg = "MISSING_OLD"
        if allow_missing_old:
            return (op, msg, None)
        raise RuntimeError(f"[row {op.rownum}] Old member not found: safe={op.safe_name} member={op.old_group}")

    # capture permissions + expiration for add/update
    permissions = old_member.get("permissions") or {}
    expiration = old_member.get("membershipExpirationDate")

    new_member = client.get_member(op.safe_name, op.new_group)
    if new_member is not None:
        if on_conflict == "skip":
            return (op, "NEW_ALREADY_EXISTS_SKIP", {"permissions": permissions, "membershipExpirationDate": expiration})
        if on_conflict == "fail":
            raise RuntimeError(f"[row {op.rownum}] New member already exists: safe={op.safe_name} member={op.new_group}")
        # update
        if dry_run:
            logging.info("[DRY] Would UPDATE new member perms: safe=%s member=%s", op.safe_name, op.new_group)
        else:
            client.update_member(op.safe_name, op.new_group, permissions, expiration)
        # optionally delete old
        if delete_old:
            if dry_run:
                logging.info("[DRY] Would DELETE old member: safe=%s member=%s", op.safe_name, op.old_group)
            else:
                client.delete_member(op.safe_name, op.old_group)
        return (op, "UPDATED_NEW_AND_DELETED_OLD" if delete_old else "UPDATED_NEW", {"permissions": permissions, "membershipExpirationDate": expiration})

    # add new
    if dry_run:
        logging.info("[DRY] Would ADD new member: safe=%s member=%s", op.safe_name, op.new_group)
    else:
        client.add_member(op.safe_name, op.new_group, permissions, search_in, member_type, expiration)

    # delete old
    if delete_old:
        if dry_run:
            logging.info("[DRY] Would DELETE old member: safe=%s member=%s", op.safe_name, op.old_group)
        else:
            client.delete_member(op.safe_name, op.old_group)

    return (op, "ADDED_NEW_AND_DELETED_OLD" if delete_old else "ADDED_NEW", {"permissions": permissions, "membershipExpirationDate": expiration})


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        prog="cyberark_safe_group_rename",
        description="Batch 'rename' CyberArk Safe member groups by add-new + (optional) delete-old.",
    )

    # Standard “info/version”
    p.add_argument("--version", action="store_true", help="Print version and exit.")
    p.add_argument("--info", action="store_true", help="Print spreadsheet summary and exit.")

    # Input
    p.add_argument("--input", required=True, help="Path to XLSX file (your export).")
    p.add_argument("--sheet-list", default="List", help="Sheet name containing rename operations. Default: List")
    p.add_argument("--environment", default=None, help="Filter operations by Environment column value.")

    # PVWA endpoint / auth
    p.add_argument("--pvwa", required=True, help="PVWA base URL, e.g. https://pvwa.company.com")
    p.add_argument("--auth-type", default="Cyberark", choices=["Cyberark", "LDAP", "Windows", "RADIUS"],
                   help="Auth type for logon endpoint. Default: Cyberark")  # :contentReference[oaicite:12]{index=12}
    p.add_argument("--username", default=os.getenv("CYBERARK_USERNAME", ""), help="API username (or env CYBERARK_USERNAME).")
    p.add_argument("--password", default=os.getenv("CYBERARK_PASSWORD", ""), help="API password (or env CYBERARK_PASSWORD).")
    p.add_argument("--token", default=os.getenv("CYBERARK_TOKEN", ""), help="Use existing token (skips logon).")
    p.add_argument("--concurrent-session", action="store_true", help="Request concurrent sessions at logon (if supported).")

    # Behavior
    p.add_argument("--search-in", default="Vault", help="Value for 'searchIn' when adding safe member (Vault or domain).")  # :contentReference[oaicite:13]{index=13}
    p.add_argument("--member-type", default="Group", choices=["Group", "User"], help="MemberType to add. Default: Group")  # :contentReference[oaicite:14]{index=14}
    p.add_argument("--on-conflict", default="update", choices=["update", "skip", "fail"],
                   help="If new member already exists: update perms, skip, or fail. Default: update")
    p.add_argument("--delete-old", action="store_true", help="Delete old safe member after add/update.")
    p.add_argument("--allow-missing-old", action="store_true", help="Skip ops where old member is missing (no error).")

    # Safety / output
    p.add_argument("--dry-run", action="store_true", help="Do not change anything; only log actions.")
    p.add_argument("--emit-curl", default="", help="Write an audit-friendly .sh script with curl commands (uses fetched perms).")

    # Perf / resilience tuning
    p.add_argument("--workers", type=int, default=8, help="Parallel workers. Default: 8")
    p.add_argument("--timeout", type=float, default=30.0, help="HTTP timeout seconds. Default: 30")
    p.add_argument("--retries", type=int, default=3, help="HTTP retries on retryable codes/timeouts. Default: 3")
    p.add_argument("--retry-backoff", type=float, default=0.5, help="Backoff base seconds. Default: 0.5")
    p.add_argument("--retry-status", default="429,500,502,503,504",
                   help="Comma list of HTTP codes to retry. Default: 429,500,502,503,504")
    p.add_argument("--pool-maxsize", type=int, default=32, help="Requests pool maxsize. Default: 32")
    p.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds before each op (rate-limit friendly). Default: 0")

    # TLS
    p.add_argument("--no-verify-tls", action="store_true", help="Disable TLS cert verification.")

    # Logging
    p.add_argument("--loglevel", default="INFO", help="DEBUG, INFO, WARNING, ERROR. Default: INFO")
    p.add_argument("--logmode", default="stdout", choices=["stdout", "file"], help="Log to stdout or file. Default: stdout")
    p.add_argument("--logfile", default="", help="Log file path (required if --logmode=file).")

    args = p.parse_args(argv)

    if args.version:
        print(__version__)
        return 0

    setup_logging(args.loglevel, args.logmode, args.logfile or None)

    ops = read_ops_from_xlsx(args.input, args.sheet_list, args.environment)

    if args.info:
        print_info(ops)
        return 0

    if not ops:
        logging.warning("No operations found (after filtering).")
        return 0

    retry_status = [int(x) for x in args.retry_status.split(",") if x.strip()]

    verify_tls = not args.no_verify_tls

    # Build a “primary” client used for logon only
    client0 = CyberArkClient(
        base_url=args.pvwa,
        token=(args.token or None),
        verify_tls=verify_tls,
        timeout_s=args.timeout,
        retries=args.retries,
        retry_backoff_s=args.retry_backoff,
        retry_status=retry_status,
        pool_maxsize=args.pool_maxsize,
    )

    token = args.token.strip() if args.token else ""
    if not token:
        if not args.username or not args.password:
            raise SystemExit("Missing credentials: provide --token OR (--username and --password).")
        token = client0.logon(args.auth_type, args.username, args.password, args.concurrent_session)

    # Each worker gets its own session, but shares the same token:
    def make_client() -> CyberArkClient:
        return CyberArkClient(
            base_url=args.pvwa,
            token=token,
            verify_tls=verify_tls,
            timeout_s=args.timeout,
            retries=args.retries,
            retry_backoff_s=args.retry_backoff,
            retry_status=retry_status,
            pool_maxsize=args.pool_maxsize,
        )

    results: List[Tuple[RenameOp, str, Optional[Dict[str, Any]]]] = []
    captured_for_curl: List[Tuple[RenameOp, Dict[str, Any]]] = []

    logging.info("Starting %d operation(s) with %d worker(s). dry_run=%s", len(ops), args.workers, args.dry_run)

    with futures.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futs = []
        for op in ops:
            c = make_client()
            futs.append(ex.submit(
                run_one_op,
                op, c,
                args.search_in,
                args.member_type,
                args.on_conflict,
                bool(args.delete_old),
                bool(args.dry_run),
                bool(args.allow_missing_old),
                float(args.sleep),
            ))
        for f in futures.as_completed(futs):
            op, status, payload = f.result()
            results.append((op, status, payload))
            if payload and args.emit_curl:
                captured_for_curl.append((op, payload))
            logging.info("Result: env=%s safe=%s %s -> %s : %s",
                         op.environment, op.safe_name, op.old_group, op.new_group, status)

    # Optionally write curl script
    if args.emit_curl:
        script = build_curl_script(
            ops_results=captured_for_curl,
            base_url=args.pvwa,
            auth_type=args.auth_type,
            username=args.username or "<username>",
            search_in=args.search_in,
            member_type=args.member_type,
            delete_old=bool(args.delete_old),
        )
        with open(args.emit_curl, "w", encoding="utf-8") as fp:
            fp.write(script)
        os.chmod(args.emit_curl, 0o750)
        logging.info("Wrote curl script: %s", args.emit_curl)

    # Logoff only if we did an actual logon in this run (i.e., no external token)
    if not args.token:
        try:
            client0.token = token
            client0.logoff()
        except Exception as e:
            logging.warning("Logoff failed: %s", e)

    # Summary
    by_status: Dict[str, int] = {}
    for _, st, _ in results:
        by_status[st] = by_status.get(st, 0) + 1
    logging.info("Summary: %s", json.dumps(by_status, indent=2, sort_keys=True))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
